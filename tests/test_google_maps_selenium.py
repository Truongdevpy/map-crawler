import unittest
import json
import sqlite3
from contextlib import closing
from pathlib import Path
from tempfile import TemporaryDirectory

import crawl_google_maps_selenium as gm


class FakeElement:
    def __init__(self, text="", **attrs):
        self.text = text
        self.attrs = attrs

    def get_attribute(self, name):
        return self.attrs.get(name, "")


class FakeDriver:
    def __init__(self, elements_by_selector):
        self.elements_by_selector = elements_by_selector

    def find_elements(self, by=None, value=None):
        return self.elements_by_selector.get(value, [])


def make_place(**overrides):
    data = {
        "name": "Khách sạn A",
        "normalized_name": "khach san a",
        "category": "khách sạn",
        "destination_id": "abc123",
        "address": "1 Đồng Khởi, Phường Bến Nghé, Quận 1, TP. Hồ Chí Minh",
        "province": "TP. Hồ Chí Minh",
        "district": "Quận 1",
        "ward": "Phường Bến Nghé",
        "description": "",
        "price_min": 100000,
        "price_max": 200000,
        "price_text": "100.000-200.000 VND",
        "rating": 4.5,
        "review_count": 123,
        "latitude": 16.1,
        "longitude": 108.2,
        "image_url": "https://example.com/a.jpg",
        "maps_url": "https://www.google.com/maps/place/a",
        "phone": "0900123456",
        "website": "https://example.com",
        "open_hours": "Mở cửa 24 giờ",
        "estimated_duration_minutes": "",
        "suitable_time": "",
        "tags": "khách sạn|Quận 1",
        "source_count": 1,
        "confidence_score": 0.9,
        "created_at": "2026-06-28T00:00:00Z",
        "updated_at": "2026-06-28T00:00:00Z",
    }
    data.update(overrides)
    return gm.Place(**data)


class GoogleMapsParserTests(unittest.TestCase):
    def test_parse_rating_accepts_decimal_comma_and_dot(self):
        self.assertEqual(gm.parse_rating("4,6"), 4.6)
        self.assertEqual(gm.parse_rating("Rating 4.7 stars"), 4.7)

    def test_parse_review_count_accepts_localized_text(self):
        self.assertEqual(gm.parse_review_count("(1.234)"), 1234)
        self.assertEqual(gm.parse_review_count("1,234 reviews"), 1234)
        self.assertEqual(gm.parse_review_count("12 bai danh gia"), 12)
        self.assertEqual(gm.parse_review_count("1,2 nghin bai danh gia"), 1200)
        self.assertEqual(gm.parse_review_count("1,2 nghìn bài đánh giá"), 1200)

    def test_clean_maps_url_removes_query_and_fragment(self):
        raw = "https://www.google.com/maps/place/Foo/@16,108,17z/data=!3m1?entry=ttu#x"
        self.assertEqual(
            gm.clean_maps_url(raw),
            "https://www.google.com/maps/place/Foo/@16,108,17z/data=!3m1",
        )

    def test_console_safe_replaces_unencodable_progress_text(self):
        self.assertEqual(gm.console_safe("Da Nang", "cp1252"), "Da Nang")
        self.assertEqual(gm.console_safe("Đà Nẵng", "cp1252"), "?à N?ng")

    def test_infer_category_from_query_handles_vietnamese_and_ascii(self):
        self.assertEqual(gm.infer_category_from_query("khách sạn Đà Nẵng"), "khách sạn")
        self.assertEqual(gm.infer_category_from_query("nha hang quan 1"), "nhà hàng")
        self.assertEqual(gm.infer_category_from_query("coffee shop Hanoi"), "quán cà phê")
        self.assertEqual(gm.infer_category_from_query("sieu thi da nang"), "siêu thị")
        self.assertEqual(gm.infer_category_from_query("ben xe mien dong"), "bến xe")

    def test_normalize_max_workers_clamps_to_safe_range(self):
        self.assertEqual(gm.normalize_max_workers(0), 1)
        self.assertEqual(gm.normalize_max_workers(2), 2)
        self.assertEqual(gm.normalize_max_workers(10), 3)

    def test_chunk_links_distributes_items_for_workers(self):
        links = [gm.ResultLink(str(index), f"https://example.com/{index}") for index in range(5)]
        chunks = gm.chunk_links(links, 2)
        self.assertEqual([[link.name_hint for link in chunk] for chunk in chunks], [["0", "2", "4"], ["1", "3"]])

    def test_all_results_limit_uses_zero_as_unlimited(self):
        links = [gm.ResultLink(str(index), f"https://example.com/{index}") for index in range(4)]

        self.assertTrue(gm.is_all_results_limit(0))
        self.assertFalse(gm.is_all_results_limit(3))
        self.assertFalse(gm.result_limit_reached(100, 0))
        self.assertTrue(gm.result_limit_reached(3, 3))
        self.assertEqual(gm.trim_result_links(links, 0), links)
        self.assertEqual(gm.trim_result_links(links, 2), links[:2])
        self.assertEqual(gm.format_limit_label(0), "ALL")

    def test_schema_fields_match_destination_table(self):
        self.assertEqual(
            gm.SCHEMA_FIELDS,
            [
                "name",
                "normalized_name",
                "category",
                "destination_id",
                "address",
                "province",
                "district",
                "ward",
                "description",
                "price_min",
                "price_max",
                "price_text",
                "rating",
                "review_count",
                "latitude",
                "longitude",
                "image_url",
                "maps_url",
                "phone",
                "website",
                "open_hours",
                "estimated_duration_minutes",
                "suitable_time",
                "tags",
                "source_count",
                "confidence_score",
                "created_at",
                "updated_at",
            ],
        )

    def test_normalize_name_removes_accents_and_noise(self):
        self.assertEqual(gm.normalize_name("  Khách sạn Đà Nẵng!!! "), "khach san da nang")

    def test_extract_district_prefers_explicit_or_address_component(self):
        self.assertEqual(
            gm.extract_district("61-63 Hoàng Kế Viêm, Ngũ Hành Sơn, Đà Nẵng 550000, Việt Nam"),
            "Ngũ Hành Sơn",
        )
        self.assertEqual(
            gm.extract_district("1 Đồng Khởi, Phường Bến Nghé, Quận 1, TP. Hồ Chí Minh"),
            "Quận 1",
        )

    def test_parse_price_range_handles_vnd_and_symbol_levels(self):
        self.assertEqual(gm.parse_price_range("100.000-250.000 ₫"), (100000, 250000))
        self.assertEqual(gm.parse_price_range("581.532 ₫ 20-21 thg 7"), (581532, 581532))
        self.assertEqual(gm.parse_price_range("₫₫"), (200000, 500000))
        self.assertEqual(gm.parse_price_range(""), (None, None))
        self.assertEqual(gm.parse_price_range("100.000-250.000 VND"), (100000, 250000))
        self.assertEqual(gm.parse_price_range("581.532 VND 20-21 thg 7"), (581532, 581532))
        self.assertEqual(gm.parse_price_range("1-100.000 \u0111"), (1, 100000))
        self.assertEqual(gm.parse_price_range("1-100.000 \u0111/ng\u01b0\u1eddi"), (1, 100000))
        self.assertEqual(gm.parse_price_range("100.000 \u0111/ng\u01b0\u1eddi"), (100000, 100000))
        self.assertEqual(gm.parse_price_range("100.000-250.000 ₫"), (100000, 250000))
        self.assertEqual(gm.parse_price_range("581.532 ₫ 20-21 thg 7"), (581532, 581532))
        self.assertEqual(gm.parse_price_range("₫₫"), (200000, 500000))
        self.assertEqual(gm.parse_price_range("4,7 (128)"), (None, None))
        self.assertEqual(gm.parse_price_range("20-21 thg 7"), (None, None))

    def test_parse_price_range_uses_category_level_fallbacks(self):
        self.assertEqual(gm.parse_price_range_for_category("₫₫", "quán cà phê"), (50_000, 150_000))
        self.assertEqual(gm.parse_price_range_for_category("Mức giá: vừa phải", "spa"), (300_000, 700_000))
        self.assertEqual(gm.parse_price_range_for_category("Price: Moderate", "quán cà phê"), (50_000, 150_000))
        self.assertEqual(gm.parse_price_range_for_category("Price: Moderate", ""), (200_000, 500_000))

    def test_extract_price_text_prefers_money_like_text(self):
        self.assertEqual(gm.extract_price_text_from_candidates(["4,7 (128)", "20-21 thg 7", "581.532 ₫"]), "581.532 ₫")
        self.assertEqual(
            gm.extract_price_text_from_candidates(["4,8 (104)", "1-100.000 \u0111/ng\u01b0\u1eddi"]),
            "1-100.000 \u0111/ng\u01b0\u1eddi",
        )
        self.assertEqual(gm.extract_price_text_from_candidates(["4,7 (128)", "Price: Moderate"]), "Price: Moderate")
        self.assertEqual(gm.extract_price_text_from_candidates(["4,7 (128)", "20-21 thg 7"]), "")

    def test_extract_coordinates_from_url_reads_at_segment(self):
        url = "https://www.google.com/maps/place/Foo/@16.0486019,108.2424468,17z/data=!4m10"
        self.assertEqual(gm.extract_coordinates_from_url(url), (16.0486019, 108.2424468))

    def test_build_destination_id_is_stable(self):
        first = gm.build_destination_id("Khách sạn A", "Đà Nẵng", 16.1, 108.2)
        second = gm.build_destination_id("Khach san A", "Da Nang", 16.1, 108.2)
        self.assertEqual(first, second)

    def test_extract_address_parts_fills_province_district_and_ward(self):
        parts = gm.extract_address_parts("1 Đồng Khởi, Phường Bến Nghé, Quận 1, TP. Hồ Chí Minh")

        self.assertEqual(parts["province"], "TP. Hồ Chí Minh")
        self.assertEqual(parts["district"], "Quận 1")
        self.assertEqual(parts["ward"], "Phường Bến Nghé")

    def test_parse_phone_number_accepts_vietnam_formats(self):
        self.assertEqual(gm.parse_phone_number("+84 90 123 4567"), "+84901234567")
        self.assertEqual(gm.parse_phone_number("Điện thoại: 0236 123 456"), "0236123456")

    def test_extract_detail_contact_fields_from_google_maps_buttons(self):
        driver = FakeDriver({
            '[data-item-id="address"] .Io6YTe': [FakeElement("123 Cầu Giấy, Hà Nội")],
            'button[data-item-id^="phone:tel:"]': [
                FakeElement("", **{"aria-label": "Số điện thoại: 024 1234 5678", "data-item-id": "phone:tel:02412345678"}),
            ],
            'a[data-item-id="authority"]': [FakeElement("", href="https://example.com")],
        })

        self.assertEqual(gm.extract_address(driver), "123 Cầu Giấy, Hà Nội")
        self.assertEqual(gm.extract_phone(driver), "02412345678")
        self.assertEqual(gm.extract_website(driver), "https://example.com")

    def test_extract_detail_contact_fields_from_aria_fallbacks(self):
        driver = FakeDriver({
            "button[aria-label], div[aria-label]": [
                FakeElement("", **{"aria-label": "Địa chỉ: 88 Trần Thái Tông, Cầu Giấy"}),
                FakeElement("", **{"aria-label": "Điện thoại: 090 123 4567"}),
                FakeElement("", **{"aria-label": "Trang web: https://spa.example.vn"}),
            ],
            'a[aria-label*="Website"]': [
                FakeElement("", href="https://maps.google.com"),
            ],
        })

        self.assertEqual(gm.extract_address(driver), "88 Trần Thái Tông, Cầu Giấy")
        self.assertEqual(gm.extract_phone(driver), "0901234567")
        self.assertEqual(gm.extract_website(driver), "https://spa.example.vn")

    def test_extract_detail_fields_from_broader_google_maps_fallbacks(self):
        driver = FakeDriver({
            '[data-item-id="address"]': [
                FakeElement("123 \u0110\u01b0\u1eddng L\u00e1ng, \u0110\u1ed1ng \u0110a, H\u00e0 N\u1ed9i"),
            ],
            "button[aria-label], div[aria-label], a[aria-label]": [
                FakeElement("", **{"aria-label": "Website: cafe-example.vn"}),
                FakeElement("", **{"aria-label": "Gi\u1edd m\u1edf c\u1eeda: M\u1edf c\u1eeda 24 gi\u1edd"}),
                FakeElement("", **{"aria-label": "104 b\u00e0i \u0111\u00e1nh gi\u00e1"}),
            ],
            'meta[property="og:image"]': [
                FakeElement("", content="https://lh3.googleusercontent.com/place-photo=w408-h306-k-no"),
            ],
        })

        self.assertEqual(gm.extract_address(driver), "123 \u0110\u01b0\u1eddng L\u00e1ng, \u0110\u1ed1ng \u0110a, H\u00e0 N\u1ed9i")
        self.assertEqual(gm.extract_website(driver), "https://cafe-example.vn")
        self.assertEqual(gm.extract_open_hours(driver), "M\u1edf c\u1eeda 24 gi\u1edd")
        self.assertEqual(gm.extract_review_count(driver), 104)
        self.assertEqual(gm.extract_image_url(driver), "https://lh3.googleusercontent.com/place-photo=w408-h306-k-no")

    def test_write_places_csv_uses_schema_or_selected_fields(self):
        place = make_place(address="Đà Nẵng", province="Đà Nẵng", district="Hải Châu", ward="")

        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "places.csv"
            gm.write_places_csv(path, [place], fieldnames=["name", "price_min", "latitude"])
            content = path.read_text(encoding="utf-8-sig").splitlines()

        self.assertEqual(content[0], "name,price_min,latitude")
        self.assertIn("Khách sạn A,100000,16.1", content[1])

    def test_write_places_csv_can_include_review_count(self):
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "reviews.csv"
            gm.write_places_csv(path, [make_place()], fieldnames=["name", "review_count"])
            content = path.read_text(encoding="utf-8-sig").splitlines()

        self.assertEqual(content[0], "name,review_count")
        self.assertTrue(content[1].endswith(",123"))

    def test_live_csv_writer_appends_rows_immediately(self):
        first = make_place(name="A", normalized_name="a", destination_id="a1", price_min=None)
        second = make_place(
            name="B",
            normalized_name="b",
            category="nhà hàng",
            destination_id="b1",
            price_min=100000,
            price_max=150000,
            rating=None,
            latitude=None,
            longitude=None,
            image_url="",
        )

        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "live.csv"
            writer = gm.LiveCsvWriter(path, ["name", "category", "price_min"])
            writer.start()
            writer.append(first)
            writer.append(second)
            content = path.read_text(encoding="utf-8-sig").splitlines()

        self.assertEqual(content[0], "name,category,price_min")
        self.assertEqual(content[1], "A,khách sạn,")
        self.assertEqual(content[2], "B,nhà hàng,100000")

    def test_write_places_jsonl_exports_one_json_object_per_line(self):
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "places.jsonl"
            gm.write_places(path, [make_place()], fieldnames=["name", "phone"], export_format="jsonl")
            rows = [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines()]

        self.assertEqual(rows, [{"name": "Khách sạn A", "phone": "0900123456"}])

    def test_write_places_sqlite_creates_places_table(self):
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "places.sqlite"
            gm.write_places(path, [make_place()], fieldnames=["name", "rating"], export_format="sqlite")
            with closing(sqlite3.connect(path)) as connection:
                with closing(connection.execute("select name, rating from places")) as cursor:
                    rows = cursor.fetchall()

        self.assertEqual(rows, [("Khách sạn A", "4.5")])

    def test_write_places_xlsx_creates_workbook(self):
        from openpyxl import load_workbook

        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "places.xlsx"
            gm.write_places(path, [make_place()], fieldnames=["name", "website"], export_format="xlsx")
            workbook = load_workbook(path)
            sheet = workbook.active

        self.assertEqual(sheet.cell(row=1, column=1).value, "name")
        self.assertEqual(sheet.cell(row=2, column=2).value, "https://example.com")

    def test_write_places_export_can_split_by_category(self):
        with TemporaryDirectory() as temp_dir:
            base = Path(temp_dir) / "places.csv"
            written = gm.write_places_export(
                base,
                [
                    make_place(name="A", category="khách sạn", destination_id="a"),
                    make_place(name="B", category="nhà hàng", destination_id="b"),
                ],
                fieldnames=["name", "category"],
                export_format="csv",
                split_by="category",
            )

        self.assertEqual(sorted(path.name for path in written), ["places_khách_sạn.csv", "places_nhà_hàng.csv"])

    def test_load_existing_keys_reads_csv_for_resume(self):
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "places.csv"
            gm.write_places_csv(path, [make_place(destination_id="abc")], fieldnames=["destination_id", "name"])

            keys = gm.load_existing_keys(path, "destination_id")

        self.assertEqual(keys, {"abc"})

    def test_build_crawl_report_counts_saved_failed_and_missing_fields(self):
        report = gm.build_crawl_report(
            saved_places=[make_place(price_min=None, image_url="", latitude=None)],
            failed_rows=[{"url": "x", "error": "boom"}],
            output_paths=[Path("data/places.csv")],
            fieldnames=["price_min", "image_url", "latitude"],
        )

        self.assertEqual(report["saved_rows"], 1)
        self.assertEqual(report["failed_rows"], 1)
        self.assertEqual(report["missing_fields"]["price_min"], 1)
        self.assertEqual(report["outputs"], ["data\\places.csv"] if "\\" in str(Path("data/places.csv")) else ["data/places.csv"])

    def test_parse_args_accepts_extended_export_and_resume_options(self):
        args = gm.parse_args([
            "khach san Da Nang",
            "--export-format",
            "jsonl",
            "--write-mode",
            "append",
            "--split-by",
            "category",
            "--resume",
            "--exclude-keywords",
            "da dong cua,tam ngung",
        ])

        self.assertEqual(args.export_format, "jsonl")
        self.assertEqual(args.write_mode, "append")
        self.assertEqual(args.split_by, "category")
        self.assertTrue(args.resume)
        self.assertEqual(args.exclude_keywords, "da dong cua,tam ngung")


    def test_parse_args_accepts_all_results_mode(self):
        args = gm.parse_args(["khach san Cau Giay", "--all-results"])

        self.assertTrue(args.all_results)

if __name__ == "__main__":
    unittest.main()
