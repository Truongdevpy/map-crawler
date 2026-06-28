#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Small Selenium crawler for Google Maps search results.

Example:
  python crawl_google_maps_selenium.py "khach san Da Nang" --limit 50 --out data/google_maps_hotels_da_nang.csv

This script does not use proxy rotation, CAPTCHA bypassing, or stealth tooling. If
Google blocks the session, slow down or use the official Places API instead.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import sqlite3
import sys
import threading
import time
import unicodedata
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Callable, Iterable, Optional
from urllib.parse import quote_plus, urlsplit, urlunsplit

import google_maps_jobs as job_utils


SCHEMA_FIELDS = [
    "name",
    "normalized_name",
    "category",
    "destination_id",
    "address",
    "province",
    "district",
    "ward",
    "description",
    "price_min",
    "price_max",
    "price_text",
    "rating",
    "review_count",
    "latitude",
    "longitude",
    "image_url",
    "maps_url",
    "phone",
    "website",
    "open_hours",
    "estimated_duration_minutes",
    "suitable_time",
    "tags",
    "source_count",
    "confidence_score",
    "created_at",
    "updated_at",
]

EXPORT_MODE_END = "end"
EXPORT_MODE_LIVE = "live"
EXPORT_MODES = (EXPORT_MODE_END, EXPORT_MODE_LIVE)

EXPORT_FORMAT_CSV = "csv"
EXPORT_FORMAT_JSONL = "jsonl"
EXPORT_FORMAT_SQLITE = "sqlite"
EXPORT_FORMAT_XLSX = "xlsx"
EXPORT_FORMATS = (EXPORT_FORMAT_CSV, EXPORT_FORMAT_JSONL, EXPORT_FORMAT_SQLITE, EXPORT_FORMAT_XLSX)

WRITE_MODE_OVERWRITE = "overwrite"
WRITE_MODE_APPEND = "append"
WRITE_MODES = (WRITE_MODE_OVERWRITE, WRITE_MODE_APPEND)

SPLIT_NONE = "none"
SPLIT_CATEGORY = "category"
SPLIT_LOCATION = "location"
SPLIT_MODES = (SPLIT_NONE, SPLIT_CATEGORY, SPLIT_LOCATION)


ALL_RESULTS_LIMIT = 0
MAX_ALL_RESULT_LINKS = 5000
STABLE_RESULT_SCROLL_ROUNDS = 8

@dataclass
class ResultLink:
    name_hint: str
    maps_url: str


@dataclass
class Place:
    name: str
    normalized_name: str
    category: str
    destination_id: str
    address: str
    province: str
    district: str
    ward: str
    description: str
    price_min: Optional[int]
    price_max: Optional[int]
    price_text: str
    rating: Optional[float]
    review_count: Optional[int]
    latitude: Optional[float]
    longitude: Optional[float]
    image_url: str
    maps_url: str
    phone: str
    website: str
    open_hours: str
    estimated_duration_minutes: str
    suitable_time: str
    tags: str
    source_count: int
    confidence_score: float
    created_at: str
    updated_at: str


@dataclass
class CrawlOptions:
    query: str
    limit: int
    out: Path
    delay: float = 1.5
    scroll_pause: float = 1.2
    timeout: float = 20
    language: str = "vi-VN"
    headless: bool = False
    user_data_dir: str | None = None
    max_workers: int = 1
    output_fields: list[str] | None = None
    export_mode: str = EXPORT_MODE_END
    export_format: str = EXPORT_FORMAT_CSV
    write_mode: str = WRITE_MODE_OVERWRITE
    split_by: str = SPLIT_NONE
    job_location: str = ""
    checkpoint_path: Path | None = None
    failed_out: Path | None = None
    report_out: Path | None = None
    resume_from_existing: bool = False
    dedupe_mode: str = "destination_id"
    exclude_keywords: list[str] | None = None
    screenshot_dir: Path | None = None
    auto_slowdown: bool = True


ProgressCallback = Callable[[str], None]


def parse_rating(text: str) -> Optional[float]:
    """Parse a Google Maps rating such as '4,6' or 'Rating 4.7 stars'."""
    if not text:
        return None

    match = re.search(r"(?<!\d)([0-5])[\.,](\d)(?!\d)", text)
    if not match:
        return None

    return float(f"{match.group(1)}.{match.group(2)}")


def parse_review_count(text: str) -> Optional[int]:
    """Parse localized review counts such as '(1.234)' or '1,2 nghin'."""
    if not text:
        return None

    normalized = (
        text.lower()
        .replace("\xa0", " ")
        .replace("(", " ")
        .replace(")", " ")
        .strip()
    )

    suffix_match = re.search(
        r"(\d+(?:[\.,]\d+)?)\s*(k|m|n|nghin|nghìn|tr|trieu|triệu)\b",
        normalized,
    )
    if suffix_match:
        number = float(suffix_match.group(1).replace(",", "."))
        suffix = suffix_match.group(2)
        multiplier = 1_000_000 if suffix in {"m", "tr", "trieu", "triệu"} else 1_000
        return int(number * multiplier)

    number_match = re.search(r"\d[\d\.,]*", normalized)
    if not number_match:
        return None

    raw = number_match.group(0)
    if re.fullmatch(r"\d{1,3}([\.,]\d{3})+", raw):
        return int(re.sub(r"[\.,]", "", raw))

    digits_only = re.sub(r"\D", "", raw)
    return int(digits_only) if digits_only else None


def parse_review_count_token(raw: str) -> Optional[int]:
    if re.fullmatch(r"\d{1,3}([\.,]\d{3})+", raw):
        return int(re.sub(r"[\.,]", "", raw))
    if re.fullmatch(r"[0-5][\.,]\d", raw):
        return None

    digits_only = re.sub(r"\D", "", raw)
    return int(digits_only) if digits_only else None


def parse_review_count(text: str) -> Optional[int]:
    """Parse review counts while ignoring rating decimals such as '4,7'."""
    if not text:
        return None

    normalized = text.lower().replace("\xa0", " ").strip()

    for raw in re.findall(r"\((\d[\d\.,]*)\)", normalized):
        count = parse_review_count_token(raw)
        if count is not None:
            return count

    suffix_match = re.search(
        r"(\d+(?:[\.,]\d+)?)\s*(k|m|n|nghin|nghìn|nghÃ¬n|tr|trieu|triệu|triá»‡u)\b",
        normalized,
    )
    if suffix_match:
        number = float(suffix_match.group(1).replace(",", "."))
        suffix = suffix_match.group(2)
        multiplier = 1_000_000 if suffix in {"m", "tr", "trieu", "triệu", "triá»‡u"} else 1_000
        return int(number * multiplier)

    for raw in re.findall(r"\d[\d\.,]*", normalized):
        count = parse_review_count_token(raw)
        if count is not None:
            return count
    return None


def clean_maps_url(url: str) -> str:
    parts = urlsplit(url)
    return urlunsplit((parts.scheme, parts.netloc, parts.path, "", ""))


def strip_label_prefix(value: str) -> str:
    return re.sub(r"^\s*(address|dia chi|dja chi|địa chỉ)\s*:\s*", "", value, flags=re.I).strip()


def strip_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFD", value)
    without_marks = "".join(char for char in normalized if unicodedata.category(char) != "Mn")
    return without_marks.replace("đ", "d").replace("Đ", "D")


def normalize_name(value: str) -> str:
    normalized = strip_accents(value).lower()
    normalized = re.sub(r"[^a-z0-9\s]+", " ", normalized)
    return re.sub(r"\s+", " ", normalized).strip()


def clean_address_component(value: str) -> str:
    value = re.sub(r"\b\d{5,6}\b", "", value)
    return re.sub(r"\s+", " ", value).strip(" ,")


def extract_district(address: str) -> str:
    if not address:
        return ""

    parts = [clean_address_component(part) for part in address.split(",")]
    parts = [part for part in parts if part and "việt nam" not in part.lower()]
    district_prefixes = (
        "quận",
        "huyện",
        "thị xã",
        "thị trấn",
    )

    for part in reversed(parts):
        lowered = part.lower()
        if lowered.startswith(district_prefixes):
            return part

    return parts[-2] if len(parts) >= 2 else ""


def extract_address_parts(address: str) -> dict[str, str]:
    if not address:
        return {"province": "", "district": "", "ward": ""}

    parts = [clean_address_component(part) for part in address.split(",")]
    parts = [part for part in parts if part and "viet nam" not in strip_accents(part).lower()]
    ward = ""
    district = ""
    province = ""

    for part in parts:
        lowered = strip_accents(part).lower()
        if not ward and lowered.startswith(("phuong", "xa", "thi tran", "ward")):
            ward = part
        if not district and lowered.startswith(("quan", "huyen", "thi xa", "thi tran", "district")):
            district = part
        if lowered.startswith(("tp", "thanh pho", "tinh", "city", "province")):
            province = part

    if not district:
        district = extract_district(address)
    if not province and parts:
        province = parts[-1]

    return {"province": province, "district": district, "ward": ward}


def parse_phone_number(text: str) -> str:
    if not text:
        return ""

    match = re.search(r"(\+?\d[\d\s().-]{7,}\d)", text)
    if not match:
        return ""

    raw = match.group(1).strip()
    prefix = "+" if raw.startswith("+") else ""
    digits = re.sub(r"\D", "", raw)
    if len(digits) < 8:
        return ""
    return f"{prefix}{digits}"


def extract_url_from_text(text: str) -> str:
    value = str(text or "").strip()
    if not value:
        return ""
    match = re.search(r"https?://[^\s,;]+", value)
    if match:
        return match.group(0).rstrip(").,;")
    if value.startswith("www."):
        return value
    return ""

def is_google_maps_url(url: str) -> bool:
    lowered = str(url or "").lower()
    return "google.com/maps" in lowered or "maps.google." in lowered

def has_price_marker(text: str) -> bool:
    normalized = text.lower().replace("\xa0", " ")
    if not has_price_marker(normalized):
        return None, None
    symbol_only_count = max(normalized.count("₫"), normalized.count("â‚«"), normalized.count("$"))
    if symbol_only_count and not re.search(r"\d", normalized):
        ranges = {
            1: (0, 200_000),
            2: (200_000, 500_000),
            3: (500_000, 1_000_000),
            4: (1_000_000, 2_000_000),
        }
        return ranges.get(min(symbol_only_count, 4), (None, None))
    normalized = normalized.replace("vnđ", "vnd").replace("₫", "vnd").replace("đ", "vnd")
    if any(marker in normalized for marker in ("₫", "â‚«", "$", "vnd", "vnđ", "dong")):
        return True
    return bool(re.search(r"\b(k|nghin|nghìn|nghÃ¬n|tr|trieu|triệu|triá»‡u)\b", normalized))


def has_price_marker(text: str) -> bool:
    normalized = text.lower().replace("\xa0", " ")
    if any(marker in normalized for marker in ("\u20ab", "â‚«", "$", "vnd", "vn\u0111", "dong")):
        return True
    return bool(re.search(r"\b(k|nghin|ngh\u00ecn|nghÃ¬n|tr|trieu|tri\u1ec7u|triá»‡u)\b", normalized))


GENERIC_PRICE_LEVEL_RANGES = {
    1: (0, 200_000),
    2: (200_000, 500_000),
    3: (500_000, 1_000_000),
    4: (1_000_000, 2_000_000),
}

PRICE_LEVEL_RANGES_BY_CATEGORY = {
    "cafe": {
        1: (0, 50_000),
        2: (50_000, 150_000),
        3: (150_000, 300_000),
        4: (300_000, 600_000),
    },
    "food": {
        1: (0, 100_000),
        2: (100_000, 300_000),
        3: (300_000, 700_000),
        4: (700_000, 1_500_000),
    },
    "spa": {
        1: (0, 300_000),
        2: (300_000, 700_000),
        3: (700_000, 1_500_000),
        4: (1_500_000, 3_000_000),
    },
    "lodging": GENERIC_PRICE_LEVEL_RANGES,
}

def parse_price_level(text: str) -> Optional[int]:
    normalized = str(text or "").lower().replace("\xa0", " ")
    symbol_count = max(normalized.count("\u20ab"), normalized.count("â‚«"), normalized.count("$"))
    if symbol_count and not re.search(r"\d", normalized):
        return min(symbol_count, 4)

    ascii_text = strip_accents(normalized)
    if not any(token in ascii_text for token in ("price", "gia", "muc gia", "cost", "chi phi")):
        return None

    if any(token in ascii_text for token in ("very expensive", "rat dat", "gia rat dat")):
        return 4
    if any(token in ascii_text for token in ("expensive", "dat", "gia cao")):
        return 3
    if any(token in ascii_text for token in ("moderate", "vua phai", "trung binh", "tam trung")):
        return 2
    if any(token in ascii_text for token in ("inexpensive", "cheap", "gia re", "binh dan")):
        return 1
    return None

def price_level_bucket_for_category(category: str) -> str:
    normalized = strip_accents(category or "").lower()
    if any(token in normalized for token in ("ca phe", "coffee", "tra sua", "tiem banh", "bakery")):
        return "cafe"
    if any(token in normalized for token in ("spa", "massage", "salon")):
        return "spa"
    if any(token in normalized for token in ("nha hang", "quan an", "quan nhau", "quan bar", "bar", "karaoke")):
        return "food"
    if any(token in normalized for token in ("khach san", "hotel", "resort", "homestay", "villa", "nha nghi", "hostel", "can ho")):
        return "lodging"
    return "generic"

def price_level_range(level: int, category: str = "") -> tuple[Optional[int], Optional[int]]:
    ranges = PRICE_LEVEL_RANGES_BY_CATEGORY.get(price_level_bucket_for_category(category), GENERIC_PRICE_LEVEL_RANGES)
    return ranges.get(min(max(int(level), 1), 4), (None, None))

def has_price_marker(text: str) -> bool:
    normalized = text.lower().replace("\xa0", " ")
    if any(marker in normalized for marker in ("\u20ab", "â‚«", "$", "vnd", "vn\u0111", "dong")):
        return True
    if parse_price_level(normalized) is not None:
        return True
    return bool(re.search(r"\b(k|nghin|ngh\u00ecn|nghÃ¬n|tr|trieu|tri\u1ec7u|triá»‡u)\b", normalized))

def extract_price_text_from_candidates(candidates: Iterable[str]) -> str:
    for candidate in candidates:
        text = re.sub(r"\s+", " ", str(candidate or "")).strip()
        if not text or not has_price_marker(text):
            continue
        price_min, price_max = parse_price_range(text)
        compact = text.replace(" ", "")
        if price_min is not None or price_max is not None or parse_price_level(text) is not None or re.fullmatch(r"[₫$â‚«]{1,4}", compact):
            return text
    return ""


def parse_price_number(value: str) -> Optional[int]:
    value = value.strip().lower().replace("\xa0", " ")
    multiplier = 1
    if re.search(r"\b(tr|triệu|trieu|m)\b", value):
        multiplier = 1_000_000
    elif re.search(r"\b(k|nghìn|nghin)\b", value):
        multiplier = 1_000

    match = re.search(r"\d+(?:[\.,]\d+)*", value)
    if not match:
        return None

    raw = match.group(0)
    if multiplier > 1 and re.fullmatch(r"\d+[\.,]\d+", raw):
        number = float(raw.replace(",", "."))
        return int(number * multiplier)

    return int(re.sub(r"\D", "", raw)) * multiplier


def parse_price_number(value: str) -> Optional[int]:
    value = value.strip().lower().replace("\xa0", " ")
    multiplier = 1
    if re.search(r"\b(tr|triệu|triá»‡u|trieu)\b", value):
        multiplier = 1_000_000
    elif re.search(r"\b(k|nghìn|nghÃ¬n|nghin)\b", value):
        multiplier = 1_000

    match = re.search(r"\d+(?:[\.,]\d+)*", value)
    if not match:
        return None

    raw = match.group(0)
    if multiplier > 1 and re.fullmatch(r"\d+[\.,]\d+", raw):
        number = float(raw.replace(",", "."))
        return int(number * multiplier)

    return int(re.sub(r"\D", "", raw)) * multiplier


def parse_price_range(text: str) -> tuple[Optional[int], Optional[int]]:
    if not text:
        return None, None

    normalized = text.lower().replace("\xa0", " ")
    price_symbol_count = max(normalized.count("₫"), normalized.count("$"))
    if price_symbol_count and not re.search(r"\d", normalized):
        ranges = {
            1: (0, 200_000),
            2: (200_000, 500_000),
            3: (500_000, 1_000_000),
            4: (1_000_000, 2_000_000),
        }
        return ranges.get(min(price_symbol_count, 4), (None, None))

    range_match = re.search(
        r"(\d+(?:[\.,]\d+)*)\s*[-–]\s*(\d+(?:[\.,]\d+)*)\s*(?:₫|vnd|đ)",
        normalized,
    )
    if range_match:
        low = parse_price_number(range_match.group(1))
        high = parse_price_number(range_match.group(2))
        if low is not None and high is not None:
            return min(low, high), max(low, high)

    currency_matches = re.findall(
        r"(?:₫|vnd|đ)\s*(\d+(?:[\.,]\d+)*)|(\d+(?:[\.,]\d+)*)\s*(?:₫|vnd|đ)",
        normalized,
    )
    currency_numbers = [
        number
        for number in (parse_price_number(left or right) for left, right in currency_matches)
        if number is not None
    ]
    if currency_numbers:
        return min(currency_numbers), max(currency_numbers)

    number_matches = re.findall(r"\d+(?:[\.,]\d+)*(?:\s*(?:k|nghìn|nghin|tr|triệu|trieu|m))", normalized)
    numbers = [number for number in (parse_price_number(match) for match in number_matches) if number is not None]
    if not numbers:
        return None, None
    if len(numbers) == 1:
        return numbers[0], numbers[0]
    return min(numbers), max(numbers)


def parse_price_range(text: str) -> tuple[Optional[int], Optional[int]]:
    if not text:
        return None, None

    normalized = text.lower().replace("\xa0", " ")
    if not has_price_marker(normalized):
        return None, None

    symbol_count = max(normalized.count("\u20ab"), normalized.count("â‚«"), normalized.count("$"))
    if symbol_count and not re.search(r"\d", normalized):
        ranges = {
            1: (0, 200_000),
            2: (200_000, 500_000),
            3: (500_000, 1_000_000),
            4: (1_000_000, 2_000_000),
        }
        return ranges.get(min(symbol_count, 4), (None, None))

    normalized = (
        normalized
        .replace("vn\u0111", "vnd")
        .replace("\u20ab", "vnd")
        .replace("\u0111", "vnd")
    )

    range_match = re.search(
        r"(\d+(?:[\.,]\d+)*)\s*[-–]\s*(\d+(?:[\.,]\d+)*)\s*(?:â‚«|vnd|\u0111)",
        normalized,
    )
    if range_match:
        low = parse_price_number(range_match.group(1))
        high = parse_price_number(range_match.group(2))
        if low is not None and high is not None:
            return min(low, high), max(low, high)

    currency_matches = re.findall(
        r"(?:â‚«|vnd|\u0111)\s*(\d+(?:[\.,]\d+)*)|(\d+(?:[\.,]\d+)*)\s*(?:â‚«|vnd|\u0111)",
        normalized,
    )
    currency_numbers = [
        number
        for number in (parse_price_number(left or right) for left, right in currency_matches)
        if number is not None
    ]
    if currency_numbers:
        return min(currency_numbers), max(currency_numbers)

    number_matches = re.findall(
        r"\d+(?:[\.,]\d+)*(?:\s*(?:k|ngh\u00ecn|nghÃ¬n|nghin|tr|tri\u1ec7u|triá»‡u|trieu))",
        normalized,
    )
    numbers = [number for number in (parse_price_number(match) for match in number_matches) if number is not None]
    if not numbers:
        return None, None
    if len(numbers) == 1:
        return numbers[0], numbers[0]
    return min(numbers), max(numbers)


def parse_price_range_for_category(text: str, category: str = "") -> tuple[Optional[int], Optional[int]]:
    level = parse_price_level(text)
    if level is not None and not re.search(r"\d", str(text or "")):
        return price_level_range(level, category)

    price_min, price_max = parse_price_range(text)
    if price_min is not None or price_max is not None:
        return price_min, price_max
    if level is not None:
        return price_level_range(level, category)
    return None, None

def extract_coordinates_from_url(url: str) -> tuple[Optional[float], Optional[float]]:
    at_match = re.search(r"@(-?\d+(?:\.\d+)?),(-?\d+(?:\.\d+)?)", url)
    if at_match:
        return float(at_match.group(1)), float(at_match.group(2))

    data_match = re.search(r"!3d(-?\d+(?:\.\d+)?)!4d(-?\d+(?:\.\d+)?)", url)
    if data_match:
        return float(data_match.group(1)), float(data_match.group(2))

    return None, None


def build_destination_id(name: str, address: str, latitude: Optional[float], longitude: Optional[float]) -> str:
    lat_lng = ""
    if latitude is not None and longitude is not None:
        lat_lng = f"{latitude:.6f},{longitude:.6f}"

    stable_key = "|".join([normalize_name(name), normalize_name(address), lat_lng])
    return hashlib.sha1(stable_key.encode("utf-8")).hexdigest()[:16]


def build_tags(category: str, district: str) -> str:
    tags = []
    for value in (category, district):
        value = value.strip()
        if value and value not in tags:
            tags.append(value)
    return "|".join(tags)


def build_confidence_score(place: Place) -> float:
    score = 0.4
    if place.name:
        score += 0.12
    if place.address:
        score += 0.12
    if place.latitude is not None and place.longitude is not None:
        score += 0.12
    if place.rating is not None:
        score += 0.05
    if place.review_count is not None:
        score += 0.03
    if place.category:
        score += 0.05
    if place.price_text:
        score += 0.02
    if place.image_url:
        score += 0.05
    if place.maps_url:
        score += 0.02
    if place.phone:
        score += 0.03
    if place.website:
        score += 0.03
    if place.open_hours:
        score += 0.03
    return round(min(score, 0.99), 2)


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def infer_category_from_query(query: str) -> str:
    normalized = re.sub(r"\s+", " ", query.lower()).strip()
    ascii_normalized = strip_accents(normalized)
    known_categories = (
        ("khách sạn", ("khách sạn", "khach san", "hotel", "hotels")),
        ("resort", ("resort",)),
        ("homestay", ("homestay",)),
        ("villa", ("villa", "biệt thự", "biet thu")),
        ("căn hộ dịch vụ", ("căn hộ dịch vụ", "can ho dich vu", "serviced apartment")),
        ("nhà nghỉ", ("nhà nghỉ", "nha nghi", "guest house", "motel")),
        ("hostel", ("hostel",)),
        ("nhà hàng", ("nhà hàng", "nha hang", "restaurant", "restaurants")),
        ("quán ăn", ("quán ăn", "quan an", "eatery")),
        ("quán cà phê", ("quán cà phê", "quán cafe", "quan ca phe", "quan cafe", "cafe", "coffee")),
        ("trà sữa", ("trà sữa", "tra sua", "milk tea")),
        ("tiệm bánh", ("tiệm bánh", "tiem banh", "bakery", "cake shop")),
        ("quán nhậu", ("quán nhậu", "quan nhau")),
        ("quán bar", ("quán bar", "quan bar", "bar", "pub")),
        ("karaoke", ("karaoke",)),
        ("spa", ("spa", "massage")),
        ("salon tóc", ("salon tóc", "salon toc", "hair salon", "barber")),
        ("gym", ("gym", "fitness")),
        ("điểm tham quan", ("điểm tham quan", "diem tham quan", "attraction", "tourist attraction")),
        ("bãi biển", ("bãi biển", "bai bien", "beach")),
        ("bảo tàng", ("bảo tàng", "bao tang", "museum")),
        ("công viên", ("công viên", "cong vien", "park")),
        ("khu vui chơi", ("khu vui chơi", "amusement", "playground")),
        ("rạp chiếu phim", ("rạp chiếu phim", "rap chieu phim", "cinema", "movie theater")),
        ("di tích lịch sử", ("di tích lịch sử", "di tich lich su", "historic site")),
        ("chùa", ("chùa", "chua", "pagoda")),
        ("nhà thờ", ("nhà thờ", "nha tho", "church")),
        ("đền", ("đền", "den", "temple")),
        ("cầu", ("cầu", "cau", "bridge")),
        ("núi", ("núi", "nui", "mountain")),
        ("thác nước", ("thác nước", "thac nuoc", "waterfall")),
        ("hồ", ("hồ ", "ho ", "lake")),
        ("đảo", ("đảo", "dao", "island")),
        ("tour du lịch", ("tour du lịch", "tour du lich", "travel tour")),
        ("công ty du lịch", ("công ty du lịch", "cong ty du lich", "travel agency")),
        ("chợ", ("chợ", "cho ", "market")),
        ("trung tâm thương mại", ("trung tâm thương mại", "trung tam thuong mai", "mall", "shopping center")),
        ("siêu thị", ("siêu thị", "sieu thi", "supermarket")),
        ("cửa hàng tiện lợi", ("cửa hàng tiện lợi", "cua hang tien loi", "convenience store")),
        ("cửa hàng lưu niệm", ("cửa hàng lưu niệm", "cua hang luu niem", "souvenir")),
        ("bệnh viện", ("bệnh viện", "benh vien", "hospital")),
        ("hiệu thuốc", ("hiệu thuốc", "hieu thuoc", "pharmacy")),
        ("phòng khám", ("phòng khám", "phong kham", "clinic")),
        ("nha khoa", ("nha khoa", "dentist", "dental")),
        ("ngân hàng", ("ngân hàng", "ngan hang", "bank")),
        ("ATM", ("atm",)),
        ("trạm xăng", ("trạm xăng", "tram xang", "gas station", "petrol station")),
        ("bãi đỗ xe", ("bãi đỗ xe", "bai do xe", "parking")),
        ("sân bay", ("sân bay", "san bay", "airport")),
        ("ga tàu", ("ga tàu", "ga tau", "train station")),
        ("bến xe", ("bến xe", "ben xe", "bus station")),
        ("bến cảng", ("bến cảng", "ben cang", "port", "harbor")),
        ("trạm xe buýt", ("trạm xe buýt", "tram xe buyt", "bus stop")),
        ("thuê xe máy", ("thuê xe máy", "thue xe may", "motorbike rental")),
        ("thuê ô tô", ("thuê ô tô", "thue o to", "car rental")),
        ("trường học", ("trường học", "truong hoc", "school")),
        ("đại học", ("đại học", "dai hoc", "university")),
        ("coworking space", ("coworking", "co-working")),
    )

    for label, tokens in known_categories:
        if any(token in normalized or token in ascii_normalized for token in tokens):
            return label
    return ""


def console_safe(value: object, encoding: str | None = None) -> str:
    target_encoding = encoding or getattr(sys.stdout, "encoding", None) or "utf-8"
    return str(value).encode(target_encoding, errors="replace").decode(target_encoding, errors="replace")


def log(message: object, *, error: bool = False) -> None:
    stream = sys.stderr if error else sys.stdout
    print(console_safe(message, getattr(stream, "encoding", None)), file=stream, flush=True)


def emit(progress: ProgressCallback | None, message: object) -> None:
    if progress is None:
        log(message)
        return
    progress(str(message))


def stop_requested(stop_event: object | None) -> bool:
    return bool(stop_event is not None and getattr(stop_event, "is_set")())


def normalize_max_workers(value: int) -> int:
    return max(1, min(3, int(value)))


def is_all_results_limit(limit: int) -> bool:
    return int(limit) == ALL_RESULTS_LIMIT

def result_limit_reached(seen_count: int, limit: int) -> bool:
    return not is_all_results_limit(limit) and seen_count >= int(limit)

def trim_result_links(links: list[ResultLink], limit: int) -> list[ResultLink]:
    if is_all_results_limit(limit):
        return list(links)
    return list(links)[: int(limit)]

def format_limit_label(limit: int) -> str:
    return "ALL" if is_all_results_limit(limit) else str(limit)

def chunk_links(links: list[ResultLink], max_workers: int) -> list[list[ResultLink]]:
    worker_count = normalize_max_workers(max_workers)
    chunks: list[list[ResultLink]] = [[] for _ in range(worker_count)]
    for index, link in enumerate(links):
        chunks[index % worker_count].append(link)
    return [chunk for chunk in chunks if chunk]


def build_driver(headless: bool, language: str, user_data_dir: str | None):
    from selenium import webdriver

    options = webdriver.ChromeOptions()
    options.add_argument(f"--lang={language}")
    options.add_argument("--window-size=1400,1000")

    if headless:
        options.add_argument("--headless=new")
    else:
        options.add_argument("--start-maximized")

    if user_data_dir:
        options.add_argument(f"--user-data-dir={user_data_dir}")

    return webdriver.Chrome(options=options)


def wait_for_body(driver, timeout: float) -> None:
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.support.ui import WebDriverWait

    WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.CSS_SELECTOR, "body")))


def accept_consent_if_present(driver) -> None:
    from selenium.webdriver.common.by import By

    wanted = ("accept all", "i agree", "agree", "toi dong y", "tôi đồng ý", "chap nhan", "chấp nhận")
    try:
        buttons = driver.find_elements(By.CSS_SELECTOR, "button")
    except Exception:
        return

    for button in buttons:
        try:
            label = " ".join(
                filter(
                    None,
                    [
                        button.text,
                        button.get_attribute("aria-label"),
                    ],
                )
            ).lower()
            if any(token in label for token in wanted):
                button.click()
                time.sleep(1)
                return
        except Exception:
            continue


def check_for_block(driver) -> None:
    from selenium.webdriver.common.by import By

    title = (driver.title or "").lower()
    try:
        body = driver.find_element(By.CSS_SELECTOR, "body").text.lower()[:3000]
    except Exception:
        body = ""

    blocked_tokens = (
        "unusual traffic",
        "our systems have detected",
        "verify you are a human",
        "not a robot",
        "captcha",
    )
    if "sorry" in title or any(token in body for token in blocked_tokens):
        raise RuntimeError("Google appears to have blocked this browser session.")


def wait_for_detail_fields(driver, timeout: float) -> None:
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait

    selectors = (
        '[data-item-id="address"]',
        '[data-item-id^="phone:tel:"]',
        '[data-item-id="authority"]',
        'button[aria-label^="Address:"]',
        'button[aria-label^="Địa chỉ:"]',
        'button[aria-label^="Phone:"]',
        'button[aria-label^="Điện thoại:"]',
        'a[aria-label*="Website"]',
        'a[aria-label*="Trang web"]',
    )
    wait_seconds = max(1.0, min(float(timeout), 4.0))
    try:
        WebDriverWait(driver, wait_seconds).until(
            lambda d: any(d.find_elements(By.CSS_SELECTOR, selector) for selector in selectors)
        )
    except Exception:
        pass

def first_text(driver, selectors: Iterable[str]) -> str:
    from selenium.webdriver.common.by import By

    for selector in selectors:
        try:
            for element in driver.find_elements(By.CSS_SELECTOR, selector):
                text = (element.text or "").strip()
                if text:
                    return text
        except Exception:
            continue
    return ""


def first_attr(driver, selectors: Iterable[str], attr: str) -> str:
    from selenium.webdriver.common.by import By

    for selector in selectors:
        try:
            for element in driver.find_elements(By.CSS_SELECTOR, selector):
                value = (element.get_attribute(attr) or "").strip()
                if value:
                    return value
        except Exception:
            continue
    return ""


def is_end_of_results_visible(driver) -> bool:
    from selenium.webdriver.common.by import By

    end_tokens = (
        "you've reached the end of the list",
        "you have reached the end of the list",
        "end of the list",
        "bạn đã xem hết danh sách",
        "ban da xem het danh sach",
        "đã xem hết danh sách",
    )
    try:
        body = driver.find_element(By.CSS_SELECTOR, "body").text.lower()
    except Exception:
        return False
    return any(token in body for token in end_tokens)

def collect_result_links(
    driver,
    query: str,
    limit: int,
    scroll_pause: float,
    timeout: float,
    progress: ProgressCallback | None = None,
    stop_event: object | None = None,
) -> list[ResultLink]:
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait

    search_url = f"https://www.google.com/maps/search/{quote_plus(query)}?hl=vi"
    driver.get(search_url)
    wait_for_body(driver, timeout)
    accept_consent_if_present(driver)
    check_for_block(driver)

    WebDriverWait(driver, timeout).until(
        lambda d: d.find_elements(By.CSS_SELECTOR, 'a[href*="/maps/place/"]')
        or "/maps/place/" in d.current_url
    )

    if "/maps/place/" in driver.current_url:
        return [ResultLink("", clean_maps_url(driver.current_url))]

    seen: dict[str, ResultLink] = {}
    stable_rounds = 0
    max_links = MAX_ALL_RESULT_LINKS if is_all_results_limit(limit) else max(1, int(limit))

    while (
        not result_limit_reached(len(seen), limit)
        and len(seen) < max_links
        and stable_rounds < STABLE_RESULT_SCROLL_ROUNDS
        and not stop_requested(stop_event)
    ):
        before = len(seen)
        anchors = driver.find_elements(By.CSS_SELECTOR, 'a[href*="/maps/place/"]')

        for anchor in anchors:
            try:
                href = anchor.get_attribute("href") or ""
                if "/maps/place/" not in href:
                    continue
                cleaned = clean_maps_url(href)
                name_hint = (anchor.get_attribute("aria-label") or anchor.text or "").splitlines()[0].strip()
                seen.setdefault(cleaned, ResultLink(name_hint=name_hint, maps_url=cleaned))
            except Exception:
                continue

        if result_limit_reached(len(seen), limit):
            break
        if is_end_of_results_visible(driver):
            emit(progress, f"End of result list detected after {len(seen)} links.")
            break

        stable_rounds = stable_rounds + 1 if len(seen) == before else 0
        scroll_result_panel(driver)
        time.sleep(scroll_pause)
        check_for_block(driver)
        emit(progress, f"Collected {len(seen)}/{format_limit_label(limit)} result links...")

    if is_all_results_limit(limit) and len(seen) >= MAX_ALL_RESULT_LINKS:
        emit(progress, f"Reached safety cap of {MAX_ALL_RESULT_LINKS} result links.")

    return trim_result_links(list(seen.values()), limit)


def scroll_result_panel(driver) -> None:
    script = """
    const feed = document.querySelector('[role="feed"]');
    if (feed) {
      feed.scrollTop = feed.scrollHeight;
      return true;
    }
    window.scrollTo(0, document.body.scrollHeight);
    return false;
    """
    try:
        driver.execute_script(script)
    except Exception:
        pass


def scrape_place(driver, result: ResultLink, timeout: float) -> Place:
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait

    driver.get(result.maps_url)
    wait_for_body(driver, timeout)
    accept_consent_if_present(driver)
    check_for_block(driver)

    WebDriverWait(driver, timeout).until(lambda d: d.find_elements(By.CSS_SELECTOR, "h1"))
    wait_for_detail_fields(driver, timeout)

    name = first_text(driver, ("h1.DUwDvf", "h1")) or result.name_hint
    category = first_text(
        driver,
        (
            'button[jsaction*="pane.rating.category"]',
            "button.DkEaL",
            'button[aria-label*="Category"]',
            'button[aria-label*="Danh mục"]',
        ),
    )

    rating = extract_rating(driver)
    review_count = extract_review_count(driver)
    address = extract_address(driver)
    description = extract_description(driver)
    price_text = extract_price_text(driver)
    price_min, price_max = parse_price_range_for_category(price_text, category)
    latitude, longitude = extract_coordinates_from_url(driver.current_url or result.maps_url)
    image_url = extract_image_url(driver)
    maps_url = clean_maps_url(driver.current_url or result.maps_url)
    phone = extract_phone(driver)
    website = extract_website(driver)
    open_hours = extract_open_hours(driver)
    suitable_time = open_hours
    address_parts = extract_address_parts(address)
    province = address_parts["province"]
    district = address_parts["district"]
    ward = address_parts["ward"]
    normalized_name = normalize_name(name)
    now = utc_now_iso()

    place = Place(
        name=name,
        normalized_name=normalized_name,
        category=category,
        destination_id=build_destination_id(name, address, latitude, longitude),
        address=address,
        province=province,
        district=district,
        ward=ward,
        description=description,
        price_min=price_min,
        price_max=price_max,
        price_text=price_text,
        rating=rating,
        review_count=review_count,
        latitude=latitude,
        longitude=longitude,
        image_url=image_url,
        maps_url=maps_url,
        phone=phone,
        website=website,
        open_hours=open_hours,
        estimated_duration_minutes="",
        suitable_time=suitable_time,
        tags=build_tags(category, district),
        source_count=1,
        confidence_score=0.0,
        created_at=now,
        updated_at=now,
    )
    place.confidence_score = build_confidence_score(place)

    return place


def extract_rating(driver) -> Optional[float]:
    from selenium.webdriver.common.by import By

    candidates: list[str] = []
    candidates.append(first_text(driver, ("div.F7nice span[aria-hidden='true']",)))
    candidates.append(first_attr(driver, ("span[role='img'][aria-label*='star']",), "aria-label"))
    candidates.append(first_attr(driver, ("span[role='img'][aria-label*='sao']",), "aria-label"))

    try:
        for element in driver.find_elements(By.CSS_SELECTOR, "div.F7nice span, div.F7nice button"):
            candidates.append(element.text or "")
            candidates.append(element.get_attribute("aria-label") or "")
    except Exception:
        pass

    for candidate in candidates:
        rating = parse_rating(candidate)
        if rating is not None:
            return rating
    return None


def extract_review_count(driver) -> Optional[int]:
    from selenium.webdriver.common.by import By

    selectors = (
        'span[aria-label*="review"]',
        'span[aria-label*="bài đánh giá"]',
        'button[aria-label*="review"]',
        'button[aria-label*="bài đánh giá"]',
        "div.F7nice span",
        "div.F7nice button",
    )
    candidates: list[str] = []

    try:
        for selector in selectors:
            for element in driver.find_elements(By.CSS_SELECTOR, selector):
                text = element.text or ""
                aria = element.get_attribute("aria-label") or ""
                combined = f"{text} {aria}".strip()
                if any(token in combined.lower() for token in ("review", "danh gia", "đánh giá", "(")):
                    candidates.append(combined)
    except Exception:
        pass

    for candidate in candidates:
        count = parse_review_count(candidate)
        if count is not None:
            return count
    return None


def extract_price_range(driver) -> tuple[Optional[int], Optional[int]]:
    from selenium.webdriver.common.by import By

    selectors = (
        "button.fT414d",
        "span.fontTitleLarge.Cbys4b",
        "div.fontLabelMedium.pUBf3e",
        "a.SlvSdc",
        "button.M77dve",
        '[data-item-id*="price"]',
        '[aria-label*="Giá"]',
        '[aria-label*="Price"]',
        "span.MW4etd + span",
        "div.F7nice span",
    )
    candidates: list[str] = []

    try:
        for selector in selectors:
            for element in driver.find_elements(By.CSS_SELECTOR, selector):
                text = " ".join(
                    part.strip()
                    for part in [element.text, element.get_attribute("aria-label")]
                    if part and part.strip()
                )
                if any(token in text for token in ("₫", "$")) or re.search(r"\b(k|nghìn|nghin|triệu|trieu)\b", text.lower()):
                    candidates.append(text)
    except Exception:
        pass

    for candidate in candidates:
        price_min, price_max = parse_price_range(candidate)
        if price_min is not None or price_max is not None:
            return price_min, price_max

    return None, None


def extract_review_count(driver) -> Optional[int]:
    from selenium.webdriver.common.by import By

    selectors = (
        'span[aria-label*="review"]',
        'span[aria-label*="reviews"]',
        'span[aria-label*="bài đánh giá"]',
        'span[aria-label*="đánh giá"]',
        'span[aria-label*="bÃ i Ä‘Ã¡nh giÃ¡"]',
        'button[aria-label*="review"]',
        'button[aria-label*="reviews"]',
        'button[aria-label*="bài đánh giá"]',
        'button[aria-label*="đánh giá"]',
        'button[aria-label*="bÃ i Ä‘Ã¡nh giÃ¡"]',
        "div.F7nice span",
        "div.F7nice button",
    )
    candidates: list[str] = []

    try:
        for selector in selectors:
            for element in driver.find_elements(By.CSS_SELECTOR, selector):
                text = element.text or ""
                aria = element.get_attribute("aria-label") or ""
                combined = f"{text} {aria}".strip()
                lowered = combined.lower()
                if any(token in lowered for token in ("review", "reviews", "danh gia", "đánh giá", "Ä‘Ã¡nh giÃ¡", "(")):
                    candidates.append(combined)
    except Exception:
        pass

    for candidate in candidates:
        count = parse_review_count(candidate)
        if count is not None:
            return count
    return None


def extract_price_text(driver) -> str:
    from selenium.webdriver.common.by import By

    selectors = (
        "button.fT414d",
        "span.fontTitleLarge.Cbys4b",
        "div.fontLabelMedium.pUBf3e",
        "a.SlvSdc",
        "button.M77dve",
        '[data-item-id*="price"]',
        '[aria-label*="Giá"]',
        '[aria-label*="GiÃ¡"]',
        '[aria-label*="giá"]',
        '[aria-label*="Mức giá"]',
        '[aria-label*="mức giá"]',
        '[aria-label*="Price"]',
        '[aria-label*="price"]',
        '[aria-label*="VND"]',
        '[aria-label*="₫"]',
    )
    candidates: list[str] = []

    try:
        for selector in selectors:
            for element in driver.find_elements(By.CSS_SELECTOR, selector):
                text = " ".join(
                    part.strip()
                    for part in [element.text, element.get_attribute("aria-label")]
                    if part and part.strip()
                )
                if has_price_marker(text):
                    candidates.append(text)
    except Exception:
        pass

    return extract_price_text_from_candidates(candidates)


def extract_price_range(driver) -> tuple[Optional[int], Optional[int]]:
    return parse_price_range(extract_price_text(driver))


def extract_image_url(driver) -> str:
    from selenium.webdriver.common.by import By

    selectors = (
        'button[aria-label*="Ảnh"] img',
        'button[aria-label*="Photo"] img',
        'img[src*="googleusercontent"]',
        'img[src*="gstatic"]',
    )
    for selector in selectors:
        try:
            for element in driver.find_elements(By.CSS_SELECTOR, selector):
                src = (element.get_attribute("src") or "").strip()
                if src and not src.startswith("data:"):
                    return src
        except Exception:
            continue

    try:
        background = driver.execute_script(
            """
            const nodes = [...document.querySelectorAll('[style*="background-image"]')];
            for (const node of nodes) {
              const style = getComputedStyle(node).backgroundImage || '';
              const match = style.match(/url\\(["']?(.*?)["']?\\)/);
              if (match && match[1] && !match[1].startsWith('data:')) return match[1];
            }
            return '';
            """
        )
        return str(background or "").strip()
    except Exception:
        return ""


def extract_description(driver) -> str:
    selectors = (
        '[data-item-id="description"]',
        "div.PYvSYb",
        "div.WeS02d",
        "div.m6QErb .fontBodyMedium",
    )
    text = first_text(driver, selectors)
    if text and len(text) > 30 and not re.search(r"^\d+(\.\d+)?$", text):
        return text
    return ""


def extract_phone(driver) -> str:
    from selenium.webdriver.common.by import By

    selectors = (
        '[data-item-id^="phone:tel:"] .Io6YTe',
        '[data-item-id^="phone:tel:"]',
        'button[data-item-id^="phone:tel:"]',
        'a[href^="tel:"]',
        'button[aria-label^="Phone:"]',
        'button[aria-label^="Điện thoại:"]',
        'button[aria-label^="Số điện thoại:"]',
        'button[aria-label*="Phone"]',
        'button[aria-label*="Điện thoại"]',
        'button[aria-label*="Số điện thoại"]',
    )
    candidates: list[str] = []
    try:
        for selector in selectors:
            for element in driver.find_elements(By.CSS_SELECTOR, selector):
                candidates.extend(
                    [
                        element.text or "",
                        element.get_attribute("aria-label") or "",
                        element.get_attribute("data-item-id") or "",
                        element.get_attribute("href") or "",
                    ]
                )
    except Exception:
        pass

    try:
        for selector in ("button[aria-label], div[aria-label]", "button[aria-label], div[aria-label], a[aria-label]"):
            for element in driver.find_elements(By.CSS_SELECTOR, selector):
                label = element.get_attribute("aria-label") or ""
                lowered = strip_accents(label).lower()
                if any(token in lowered for token in ("phone", "dien thoai", "so dien thoai")):
                    candidates.append(label)
    except Exception:
        pass

    for candidate in candidates:
        phone = parse_phone_number(candidate)
        if phone:
            return phone
    return ""


def extract_website(driver) -> str:
    from selenium.webdriver.common.by import By

    selectors = (
        'a[data-item-id="authority"]',
        '[data-item-id="authority"] a',
        '[data-item-id="authority"]',
        'a[aria-label*="Website"]',
        'a[aria-label*="Trang web"]',
        'a[aria-label*="website"]',
        'a[aria-label*="trang web"]',
    )
    candidates: list[str] = []
    try:
        for selector in selectors:
            for element in driver.find_elements(By.CSS_SELECTOR, selector):
                candidates.extend(
                    [
                        element.get_attribute("href") or "",
                        element.get_attribute("aria-label") or "",
                        element.text or "",
                    ]
                )
    except Exception:
        pass

    try:
        for selector in ("button[aria-label], div[aria-label]", "button[aria-label], div[aria-label], a[aria-label]"):
            for element in driver.find_elements(By.CSS_SELECTOR, selector):
                label = element.get_attribute("aria-label") or ""
                lowered = strip_accents(label).lower()
                if any(token in lowered for token in ("website", "trang web")):
                    candidates.append(label)
    except Exception:
        pass

    for candidate in candidates:
        website = extract_url_from_text(candidate)
        if website and not is_google_maps_url(website):
            return website
    return ""


def extract_open_hours(driver) -> str:
    return extract_suitable_time(driver)


def extract_suitable_time(driver) -> str:
    from selenium.webdriver.common.by import By

    selectors = (
        '[data-item-id*="oh"]',
        '[aria-label*="Giờ mở cửa"]',
        '[aria-label*="Hours"]',
        '[aria-label*="Open"]',
    )
    try:
        for selector in selectors:
            for element in driver.find_elements(By.CSS_SELECTOR, selector):
                text = " ".join(
                    part.strip()
                    for part in [element.text, element.get_attribute("aria-label")]
                    if part and part.strip()
                )
                if any(token in text.lower() for token in ("mở cửa", "đóng cửa", "open", "closed", "24")):
                    return re.sub(r"\s+", " ", text).strip()
    except Exception:
        pass
    return ""


def extract_address(driver) -> str:
    from selenium.webdriver.common.by import By

    address = first_text(driver, ('[data-item-id="address"] .Io6YTe',))
    if address:
        return address

    aria_address = first_attr(
        driver,
        (
            '[data-item-id="address"]',
            'button[aria-label^="Address:"]',
            'button[aria-label^="Địa chỉ:"]',
        ),
        "aria-label",
    )
    if aria_address:
        return strip_label_prefix(aria_address)

    try:
        for element in driver.find_elements(By.CSS_SELECTOR, "button[aria-label], div[aria-label]"):
            label = element.get_attribute("aria-label") or ""
            if label.lower().startswith(("address:", "địa chỉ:")):
                return strip_label_prefix(label)
    except Exception:
        pass

    return ""


def resolve_output_fields(fieldnames: list[str] | None = None) -> list[str]:
    output_fields = fieldnames or SCHEMA_FIELDS
    unknown_fields = [field for field in output_fields if field not in SCHEMA_FIELDS]
    if unknown_fields:
        raise ValueError(f"Unknown output fields: {', '.join(unknown_fields)}")
    return output_fields


def place_to_csv_row(place: Place, fieldnames: list[str]) -> dict[str, object]:
    row = asdict(place)
    return {field: row.get(field, "") for field in fieldnames}


def write_places_csv(path: Path, places: Iterable[Place], fieldnames: list[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    output_fields = resolve_output_fields(fieldnames)

    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=output_fields)
        writer.writeheader()
        for place in places:
            writer.writerow(place_to_csv_row(place, output_fields))


def infer_export_format(path: Path, export_format: str | None = None) -> str:
    if export_format:
        normalized = export_format.lower().strip()
    else:
        suffix = path.suffix.lower()
        normalized = {
            ".csv": EXPORT_FORMAT_CSV,
            ".jsonl": EXPORT_FORMAT_JSONL,
            ".sqlite": EXPORT_FORMAT_SQLITE,
            ".db": EXPORT_FORMAT_SQLITE,
            ".xlsx": EXPORT_FORMAT_XLSX,
        }.get(suffix, EXPORT_FORMAT_CSV)
    if normalized not in EXPORT_FORMATS:
        raise ValueError(f"export_format must be one of: {', '.join(EXPORT_FORMATS)}")
    return normalized


def normalize_write_mode(write_mode: str) -> str:
    normalized = (write_mode or WRITE_MODE_OVERWRITE).lower().strip()
    if normalized not in WRITE_MODES:
        raise ValueError(f"write_mode must be one of: {', '.join(WRITE_MODES)}")
    return normalized


def place_to_export_row(place: Place, fieldnames: list[str]) -> dict[str, object]:
    row = asdict(place)
    return {field: row.get(field, "") for field in fieldnames}


def write_places(
    path: Path,
    places: Iterable[Place],
    fieldnames: list[str] | None = None,
    export_format: str | None = None,
    write_mode: str = WRITE_MODE_OVERWRITE,
) -> None:
    path = Path(path)
    output_fields = resolve_output_fields(fieldnames)
    normalized_format = infer_export_format(path, export_format)
    normalized_mode = normalize_write_mode(write_mode)
    rows = [place_to_export_row(place, output_fields) for place in places]

    if normalized_format == EXPORT_FORMAT_CSV:
        write_rows_csv(path, rows, output_fields, append=normalized_mode == WRITE_MODE_APPEND)
    elif normalized_format == EXPORT_FORMAT_JSONL:
        write_rows_jsonl(path, rows, append=normalized_mode == WRITE_MODE_APPEND)
    elif normalized_format == EXPORT_FORMAT_SQLITE:
        write_rows_sqlite(path, rows, output_fields, append=normalized_mode == WRITE_MODE_APPEND)
    elif normalized_format == EXPORT_FORMAT_XLSX:
        write_rows_xlsx(path, rows, output_fields, append=normalized_mode == WRITE_MODE_APPEND)


def write_rows_csv(path: Path, rows: list[dict[str, object]], fieldnames: list[str], append: bool = False) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    needs_header = not append or not path.exists() or path.stat().st_size == 0
    mode = "a" if append else "w"
    with path.open(mode, newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        if needs_header:
            writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_rows_jsonl(path: Path, rows: list[dict[str, object]], append: bool = False) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    mode = "a" if append else "w"
    with path.open(mode, encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")


def write_rows_sqlite(path: Path, rows: list[dict[str, object]], fieldnames: list[str], append: bool = False) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    connection = sqlite3.connect(path)
    try:
        quoted_fields = ", ".join(f'"{field}" TEXT' for field in fieldnames)
        if not append:
            connection.execute("DROP TABLE IF EXISTS places")
        connection.execute(f"CREATE TABLE IF NOT EXISTS places ({quoted_fields})")
        placeholders = ", ".join("?" for _ in fieldnames)
        columns = ", ".join(f'"{field}"' for field in fieldnames)
        values = [
            ["" if row.get(field) is None else str(row.get(field, "")) for field in fieldnames]
            for row in rows
        ]
        if values:
            connection.executemany(f"INSERT INTO places ({columns}) VALUES ({placeholders})", values)
        connection.commit()
    finally:
        connection.close()


def write_rows_xlsx(path: Path, rows: list[dict[str, object]], fieldnames: list[str], append: bool = False) -> None:
    from openpyxl import Workbook, load_workbook

    path.parent.mkdir(parents=True, exist_ok=True)
    if append and path.exists():
        workbook = load_workbook(path)
        sheet = workbook.active
        if sheet.max_row == 0:
            sheet.append(fieldnames)
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "places"
        sheet.append(fieldnames)

    for row in rows:
        sheet.append(["" if row.get(field) is None else row.get(field, "") for field in fieldnames])
    workbook.save(path)


def write_places_export(
    path: Path,
    places: Iterable[Place],
    fieldnames: list[str] | None = None,
    export_format: str | None = None,
    write_mode: str = WRITE_MODE_OVERWRITE,
    split_by: str = SPLIT_NONE,
    default_location: str = "",
) -> list[Path]:
    path = Path(path)
    output_fields = resolve_output_fields(fieldnames)
    normalized_format = infer_export_format(path, export_format)
    normalized_split = (split_by or SPLIT_NONE).lower().strip()
    if normalized_split not in SPLIT_MODES:
        raise ValueError(f"split_by must be one of: {', '.join(SPLIT_MODES)}")

    places_list = list(places)
    if normalized_split == SPLIT_NONE:
        write_places(path, places_list, output_fields, normalized_format, write_mode)
        return [path]

    groups: dict[str, list[Place]] = {}
    for place in places_list:
        if normalized_split == SPLIT_CATEGORY:
            group_name = place.category or "unknown"
        else:
            group_name = place.province or place.district or default_location or "unknown"
        groups.setdefault(group_name, []).append(place)

    written: list[Path] = []
    for group_name, grouped_places in groups.items():
        group_path = path.with_name(f"{path.stem}_{job_utils.slugify_filename(group_name)}{path.suffix}")
        write_places(group_path, grouped_places, output_fields, normalized_format, write_mode)
        written.append(group_path)
    return written


def rows_from_csv(path: Path) -> list[dict[str, object]]:
    if not path.exists():
        return []
    with path.open("r", newline="", encoding="utf-8-sig") as f:
        return [dict(row) for row in csv.DictReader(f)]


def load_existing_keys(path: Path, dedupe_mode: str = "destination_id") -> set[str]:
    rows = rows_from_csv(Path(path))
    return {key for key in (job_utils.dedupe_key(row, dedupe_mode) for row in rows) if key}


def place_matches_existing(place: Place, existing_keys: set[str], dedupe_mode: str) -> bool:
    if not existing_keys:
        return False
    key = job_utils.dedupe_key(place_to_export_row(place, SCHEMA_FIELDS), dedupe_mode)
    return bool(key and key in existing_keys)


def write_failed_rows(path: Path, failed_rows: list[dict[str, object]]) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = ["query", "index", "name_hint", "url", "error", "screenshot", "created_at"]
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in failed_rows:
            writer.writerow({field: row.get(field, "") for field in fieldnames})


def build_crawl_report(
    saved_places: list[Place],
    failed_rows: list[dict[str, object]],
    output_paths: list[Path],
    fieldnames: list[str] | None = None,
) -> dict[str, object]:
    output_fields = resolve_output_fields(fieldnames)
    rows = [place_to_export_row(place, output_fields) for place in saved_places]
    missing_report = job_utils.build_missing_field_report(rows, output_fields)
    return {
        "saved_rows": len(saved_places),
        "failed_rows": len(failed_rows),
        "outputs": [str(path) for path in output_paths],
        "missing_fields": missing_report["missing"],
        "created_at": utc_now_iso(),
    }


def write_crawl_report(path: Path, report: dict[str, object]) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")


class LiveCsvWriter:
    def __init__(self, path: Path, fieldnames: list[str] | None = None) -> None:
        self.path = path
        self.fieldnames = resolve_output_fields(fieldnames)
        self._lock = threading.Lock()
        self._started = False

    def start(self) -> None:
        with self._lock:
            self.path.parent.mkdir(parents=True, exist_ok=True)
            with self.path.open("w", newline="", encoding="utf-8-sig") as f:
                writer = csv.DictWriter(f, fieldnames=self.fieldnames)
                writer.writeheader()
            self._started = True

    def append(self, place: Place) -> None:
        with self._lock:
            if not self._started:
                self.path.parent.mkdir(parents=True, exist_ok=True)
                with self.path.open("w", newline="", encoding="utf-8-sig") as f:
                    writer = csv.DictWriter(f, fieldnames=self.fieldnames)
                    writer.writeheader()
                self._started = True

            with self.path.open("a", newline="", encoding="utf-8-sig") as f:
                writer = csv.DictWriter(f, fieldnames=self.fieldnames)
                writer.writerow(place_to_csv_row(place, self.fieldnames))


def wait_if_paused(pause_event: object | None, stop_event: object | None, progress: ProgressCallback | None = None) -> None:
    if pause_event is None:
        return
    announced = False
    while bool(getattr(pause_event, "is_set")()) and not stop_requested(stop_event):
        if not announced:
            emit(progress, "Paused. Waiting to resume...")
            announced = True
        time.sleep(0.3)


def is_block_error(exc: Exception) -> bool:
    text = str(exc).lower()
    return any(token in text for token in ("blocked", "captcha", "unusual traffic", "not a robot"))


def default_sidecar_path(out: Path, suffix: str) -> Path:
    return out.with_name(f"{out.stem}_{suffix}")


def screenshot_path_for(options: CrawlOptions, index: int) -> Path:
    folder = options.screenshot_dir or (options.out.parent / "screenshots")
    folder.mkdir(parents=True, exist_ok=True)
    return folder / f"error_{index}_{datetime.now():%Y%m%d_%H%M%S}.png"


def take_error_screenshot(driver, options: CrawlOptions, index: int) -> str:
    try:
        path = screenshot_path_for(options, index)
        driver.save_screenshot(str(path))
        return str(path)
    except Exception:
        return ""


def save_runtime_checkpoint(
    options: CrawlOptions,
    completed_urls: set[str],
    failed_rows: list[dict[str, object]],
    saved_count: int,
) -> None:
    checkpoint_path = options.checkpoint_path or default_sidecar_path(options.out, "checkpoint.json")
    job_utils.save_checkpoint(
        checkpoint_path,
        completed_ids=completed_urls,
        failed_rows=failed_rows,
        extra={"saved_rows": saved_count, "query": options.query},
    )


def scrape_place_batch(
    options: CrawlOptions,
    indexed_links: list[tuple[int, int, ResultLink]],
    fallback_category: str,
    progress: ProgressCallback | None,
    stop_event: object | None,
    pause_event: object | None = None,
    driver=None,
    on_place: Callable[[int, Place], None] | None = None,
    on_failure: Callable[[dict[str, object]], None] | None = None,
    should_skip_place: Callable[[Place], bool] | None = None,
) -> list[tuple[int, Place]]:
    owns_driver = driver is None
    if driver is None:
        driver = build_driver(options.headless, options.language, options.user_data_dir)

    scraped: list[tuple[int, Place]] = []
    current_delay = options.delay
    error_count = 0

    try:
        for index, total, result in indexed_links:
            if stop_requested(stop_event):
                emit(progress, "Stop requested. Finishing current cleanup...")
                break
            wait_if_paused(pause_event, stop_event, progress)
            if stop_requested(stop_event):
                break

            label = result.name_hint or result.maps_url
            emit(progress, f"[{index}/{total}] Scraping {label}")

            try:
                place = scrape_place(driver, result, options.timeout)
                if not place.category and fallback_category:
                    place.category = fallback_category
                    place.tags = build_tags(place.category, place.district)
                    place.confidence_score = build_confidence_score(place)
                if options.exclude_keywords and job_utils.should_exclude_row(
                    place_to_export_row(place, SCHEMA_FIELDS),
                    options.exclude_keywords,
                ):
                    emit(progress, f"  Excluded by keyword: {place.name}")
                    continue
                if should_skip_place is not None and should_skip_place(place):
                    emit(progress, f"  Skipped duplicate: {place.name}")
                    continue
                scraped.append((index, place))
                if on_place is not None:
                    on_place(index, place)
            except Exception as exc:
                error_count += 1
                screenshot = take_error_screenshot(driver, options, index)
                failed_row = {
                    "query": options.query,
                    "index": index,
                    "name_hint": result.name_hint,
                    "url": result.maps_url,
                    "error": str(exc),
                    "screenshot": screenshot,
                    "created_at": utc_now_iso(),
                }
                if on_failure is not None:
                    on_failure(failed_row)
                emit(progress, f"  Skipped: {exc}")
                if screenshot:
                    emit(progress, f"  Screenshot: {screenshot}")
                if is_block_error(exc):
                    emit(progress, "Google block/CAPTCHA detected. Soft stopping crawl.")
                    if stop_event is not None and hasattr(stop_event, "set"):
                        stop_event.set()
                    break
                if options.auto_slowdown and error_count % 3 == 0:
                    current_delay = min(max(current_delay, 1.0) * 1.5, 30.0)
                    emit(progress, f"Many errors detected. Delay increased to {current_delay:.1f}s")

            if current_delay > 0 and not stop_requested(stop_event):
                time.sleep(current_delay)
    finally:
        if owns_driver:
            driver.quit()

    return scraped


def run_crawl(
    options: CrawlOptions,
    progress: ProgressCallback | None = None,
    stop_event: object | None = None,
    pause_event: object | None = None,
) -> list[Place]:
    if options.limit < 0:
        raise ValueError("limit must be 0 for all results or at least 1")
    if options.export_mode not in EXPORT_MODES:
        raise ValueError(f"export_mode must be one of: {', '.join(EXPORT_MODES)}")
    infer_export_format(options.out, options.export_format)
    normalize_write_mode(options.write_mode)
    if options.split_by not in SPLIT_MODES:
        raise ValueError(f"split_by must be one of: {', '.join(SPLIT_MODES)}")

    options.max_workers = normalize_max_workers(options.max_workers)
    output_fields = resolve_output_fields(options.output_fields)
    fallback_category = infer_category_from_query(options.query)
    driver = build_driver(options.headless, options.language, options.user_data_dir)
    places_with_index: list[tuple[int, Place]] = []
    output_paths: list[Path] = []
    output_path_set: set[str] = set()
    failed_rows: list[dict[str, object]] = []
    checkpoint = job_utils.load_checkpoint(options.checkpoint_path or default_sidecar_path(options.out, "checkpoint.json"))
    completed_urls = set(str(url) for url in checkpoint.get("completed_ids", []) if str(url))
    if options.resume_from_existing:
        failed_rows.extend(dict(row) for row in checkpoint.get("failed_rows", []))
    existing_keys = load_existing_keys(options.out, options.dedupe_mode) if options.resume_from_existing and options.out.exists() else set()
    seen_keys = set(existing_keys)
    runtime_lock = threading.Lock()
    live_started = False

    def remember_output(paths: list[Path]) -> None:
        for path in paths:
            key = str(path)
            if key not in output_path_set:
                output_path_set.add(key)
                output_paths.append(path)

    def should_skip_place(place: Place) -> bool:
        key = job_utils.dedupe_key(place_to_export_row(place, SCHEMA_FIELDS), options.dedupe_mode)
        if not key:
            return False
        with runtime_lock:
            if key in seen_keys:
                return True
            seen_keys.add(key)
        return False

    def live_append(index: int, place: Place) -> None:
        nonlocal live_started
        with runtime_lock:
            if options.export_mode == EXPORT_MODE_LIVE:
                if not live_started:
                    remember_output(
                        write_places_export(
                            options.out,
                            [],
                            output_fields,
                            options.export_format,
                            options.write_mode,
                            options.split_by,
                            options.job_location,
                        )
                    )
                    live_started = True
                    emit(progress, f"Live export enabled: writing rows to {options.out}")
                remember_output(
                    write_places_export(
                        options.out,
                        [place],
                        output_fields,
                        options.export_format,
                        WRITE_MODE_APPEND,
                        options.split_by,
                        options.job_location,
                    )
                )
                emit(progress, f"Live exported row {index} to {options.out}")
            save_runtime_checkpoint(options, completed_urls, failed_rows, len(seen_keys) - len(existing_keys))

    def record_failure(row: dict[str, object]) -> None:
        with runtime_lock:
            failed_rows.append(row)
            save_runtime_checkpoint(options, completed_urls, failed_rows, len(seen_keys) - len(existing_keys))

    try:
        links = collect_result_links(
            driver,
            options.query,
            options.limit,
            options.scroll_pause,
            options.timeout,
            progress=progress,
            stop_event=stop_event,
        )
        emit(progress, f"Found {len(links)} result links.")

        if options.resume_from_existing and completed_urls:
            before_count = len(links)
            links = [link for link in links if link.maps_url not in completed_urls]
            skipped_count = before_count - len(links)
            if skipped_count:
                emit(progress, f"Resume skipped {skipped_count} checkpoint URLs.")

        if stop_requested(stop_event) or not links:
            remember_output(
                write_places_export(
                    options.out,
                    [],
                    output_fields,
                    options.export_format,
                    options.write_mode,
                    options.split_by,
                    options.job_location,
                )
            )
            failed_path = options.failed_out or default_sidecar_path(options.out, "failed_rows.csv")
            report_path = options.report_out or default_sidecar_path(options.out, "crawl_report.json")
            write_failed_rows(failed_path, failed_rows)
            report = build_crawl_report([], failed_rows, output_paths, output_fields)
            write_crawl_report(report_path, report)
            emit(progress, f"Saved 0 places to {options.out}")
            return []

        indexed_links = [(index, len(links), result) for index, result in enumerate(links, start=1)]
        index_to_url = {index: result.maps_url for index, _, result in indexed_links}

        def checkpoint_place(index: int, place: Place) -> None:
            url = index_to_url.get(index)
            if url:
                completed_urls.add(url)
            live_append(index, place)

        if options.max_workers == 1:
            places_with_index = scrape_place_batch(
                options,
                indexed_links,
                fallback_category,
                progress,
                stop_event,
                pause_event= pause_event,
                driver=driver,
                on_place=checkpoint_place,
                on_failure=record_failure,
                should_skip_place=should_skip_place,
            )
        else:
            driver.quit()
            driver = None
            link_chunks = chunk_links(links, options.max_workers)
            indexed_chunks: list[list[tuple[int, int, ResultLink]]] = []
            for chunk in link_chunks:
                indexed_chunks.append(
                    [(links.index(result) + 1, len(links), result) for result in chunk]
                )

            with ThreadPoolExecutor(max_workers=options.max_workers) as executor:
                futures = [
                    executor.submit(
                        scrape_place_batch,
                        options,
                        chunk,
                        fallback_category,
                        progress,
                        stop_event,
                        pause_event,
                        None,
                        checkpoint_place,
                        record_failure,
                        should_skip_place,
                    )
                    for chunk in indexed_chunks
                ]
                for future in as_completed(futures):
                    places_with_index.extend(future.result())

        places = [place for _, place in sorted(places_with_index, key=lambda item: item[0])]
        if options.export_mode == EXPORT_MODE_END:
            remember_output(
                write_places_export(
                    options.out,
                    places,
                    output_fields,
                    options.export_format,
                    options.write_mode,
                    options.split_by,
                    options.job_location,
                )
            )
        elif not live_started:
            remember_output(
                write_places_export(
                    options.out,
                    [],
                    output_fields,
                    options.export_format,
                    options.write_mode,
                    options.split_by,
                    options.job_location,
                )
            )

        failed_path = options.failed_out or default_sidecar_path(options.out, "failed_rows.csv")
        report_path = options.report_out or default_sidecar_path(options.out, "crawl_report.json")
        write_failed_rows(failed_path, failed_rows)
        report = build_crawl_report(places, failed_rows, output_paths or [options.out], output_fields)
        write_crawl_report(report_path, report)
        save_runtime_checkpoint(options, completed_urls, failed_rows, len(places))
        emit(progress, f"Saved {len(places)} places to {options.out}")
        if failed_rows:
            emit(progress, f"Saved {len(failed_rows)} failed rows to {failed_path}")
        emit(progress, f"Saved crawl report to {report_path}")
        return places
    finally:
        if driver is not None:
            driver.quit()


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Crawl basic Google Maps search result data with Selenium.")
    parser.add_argument("query", help="Search keyword, for example: \"khach san Da Nang\"")
    parser.add_argument("--limit", type=int, default=50, help="Maximum number of places to export. Use 0 for all results found.")
    parser.add_argument("--all-results", action="store_true", help="Scroll until Google Maps stops returning new places.")
    parser.add_argument("--out", type=Path, default=Path("data/google_maps.csv"), help="Output CSV path.")
    parser.add_argument("--delay", type=float, default=1.5, help="Delay between place pages, in seconds.")
    parser.add_argument("--scroll-pause", type=float, default=1.2, help="Delay between result-list scrolls, in seconds.")
    parser.add_argument("--timeout", type=float, default=20, help="Selenium wait timeout, in seconds.")
    parser.add_argument("--language", default="vi-VN", help="Browser language, default: vi-VN.")
    parser.add_argument("--headless", action="store_true", help="Run Chrome in headless mode.")
    parser.add_argument("--user-data-dir", default=None, help="Optional Chrome user data directory.")
    parser.add_argument("--max-workers", type=int, default=1, help="Parallel Chrome workers, clamped to 1-3.")
    parser.add_argument(
        "--export-mode",
        choices=EXPORT_MODES,
        default=EXPORT_MODE_END,
        help="Export mode: 'end' writes after crawling, 'live' appends rows while crawling.",
    )
    parser.add_argument(
        "--export-format",
        choices=EXPORT_FORMATS,
        default=EXPORT_FORMAT_CSV,
        help="Output format: csv, jsonl, sqlite, or xlsx.",
    )
    parser.add_argument(
        "--write-mode",
        choices=WRITE_MODES,
        default=WRITE_MODE_OVERWRITE,
        help="Use overwrite for a fresh file or append to merge into an existing output.",
    )
    parser.add_argument(
        "--split-by",
        choices=SPLIT_MODES,
        default=SPLIT_NONE,
        help="Split output into separate files by category or location.",
    )
    parser.add_argument("--job-location", default="", help="Location label used when splitting by location.")
    parser.add_argument("--checkpoint", type=Path, default=None, help="Checkpoint JSON path.")
    parser.add_argument("--failed-out", type=Path, default=None, help="Failed rows CSV path.")
    parser.add_argument("--report-out", type=Path, default=None, help="Crawl report JSON path.")
    parser.add_argument("--resume", action="store_true", help="Resume from checkpoint/output and skip known rows.")
    parser.add_argument(
        "--dedupe-mode",
        choices=("destination_id", "name_address", "coordinates"),
        default="destination_id",
        help="Key used for duplicate detection when resuming or merging.",
    )
    parser.add_argument(
        "--exclude-keywords",
        default="",
        help="Comma/semicolon/newline separated keywords to drop after scraping, e.g. closed places.",
    )
    parser.add_argument("--screenshot-dir", type=Path, default=None, help="Folder for error screenshots.")
    parser.add_argument(
        "--no-auto-slowdown",
        action="store_true",
        help="Disable automatic delay increase after repeated scrape errors.",
    )
    parser.add_argument(
        "--fields",
        default=",".join(SCHEMA_FIELDS),
        help="Comma-separated CSV fields. Defaults to the full destination schema.",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv or sys.argv[1:])
    limit = ALL_RESULTS_LIMIT if args.all_results else args.limit
    if limit < 0:
        log("--limit must be 0 for all results or at least 1", error=True)
        return 2

    options = CrawlOptions(
        query=args.query,
        limit=limit,
        out=args.out,
        delay=args.delay,
        scroll_pause=args.scroll_pause,
        timeout=args.timeout,
        language=args.language,
        headless=args.headless,
        user_data_dir=args.user_data_dir,
        max_workers=args.max_workers,
        output_fields=[field.strip() for field in args.fields.split(",") if field.strip()],
        export_mode=args.export_mode,
        export_format=args.export_format,
        write_mode=args.write_mode,
        split_by=args.split_by,
        job_location=args.job_location,
        checkpoint_path=args.checkpoint,
        failed_out=args.failed_out,
        report_out=args.report_out,
        resume_from_existing=args.resume,
        dedupe_mode=args.dedupe_mode,
        exclude_keywords=job_utils.split_multi_value(args.exclude_keywords),
        screenshot_dir=args.screenshot_dir,
        auto_slowdown=not args.no_auto_slowdown,
    )
    run_crawl(options)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
